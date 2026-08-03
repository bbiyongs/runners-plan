import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    wb = openpyxl.Workbook()
    
    # Remove default sheet to build custom order
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Fonts
    font_name = "Malgun Gothic"
    title_font = Font(name=font_name, size=14, bold=True, color="1A365D")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=9.5, color="2D3748")
    
    # Fills
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Priority colors (soft background with matching dark text)
    fill_essential = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid") # 필수 (soft red)
    font_essential = Font(name=font_name, size=9.5, bold=True, color="9B2C2C")
    
    fill_normal = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid") # 보통 (soft yellow)
    font_normal = Font(name=font_name, size=9.5, bold=True, color="975A16")
    
    fill_low = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # 낮음 (soft gray)
    font_low = Font(name=font_name, size=9.5, bold=True, color="4A5568")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left_pad = Alignment(horizontal="left", vertical="center", indent=1)
    
    # Borders
    thin_border_side = Side(style='thin', color='CBD5E0')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    bottom_heavy_border = Border(bottom=Side(style='medium', color='1A365D'))

    # ==========================================
    # 1. Tab: 버전관리 (Version History)
    # ==========================================
    ws_ver = wb.create_sheet("버전관리")
    ws_ver.views.sheetView[0].showGridLines = True
    
    # Column Widths
    ws_ver.column_dimensions['A'].width = 3
    ws_ver.column_dimensions['B'].width = 12  # 버전
    ws_ver.column_dimensions['C'].width = 16  # 변경일자
    ws_ver.column_dimensions['D'].width = 48  # 변경 내용
    ws_ver.column_dimensions['E'].width = 22  # 작성자
    ws_ver.column_dimensions['F'].width = 15  # 비고
    
    # Title
    ws_ver.row_dimensions[2].height = 36
    ws_ver.merge_cells("B2:F2")
    t_cell = ws_ver["B2"]
    t_cell.value = "요구사항 명세서 - 버전 관리 이력"
    t_cell.font = title_font
    t_cell.alignment = align_left_pad
    for col in range(2, 7):
        ws_ver.cell(row=2, column=col).border = bottom_heavy_border
        
    # Table Headers
    ws_ver.row_dimensions[4].height = 24
    ver_headers = ["버전", "변경 일자", "변경 내용", "작성자", "비고"]
    for idx, h in enumerate(ver_headers, 2):
        c = ws_ver.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = thin_border
        
    # Version Data Rows
    ver_records = [
        ("v1.0", "2026-06-18", "최초 요구사항 정의서 작성 및 멀티 탭 문서화", "Running Coach 개발팀", "-"),
        ("v1.1", "2026-06-22", "산출물 가이드(5-10p) 요구 사양 기준 개정 (Usecase, FR, NFR 분리)", "Running Coach 기획팀", "-")
    ]
    for r_idx, record in enumerate(ver_records, 5):
        ws_ver.row_dimensions[r_idx].height = 20
        row_fill = zebra_fill if r_idx % 2 == 0 else white_fill
        for c_idx, val in enumerate(record, 2):
            c = ws_ver.cell(row=r_idx, column=c_idx, value=val)
            c.font = data_font
            c.fill = row_fill
            c.border = thin_border
            if c_idx in (2, 3, 5):  # Version, Date, Author
                c.alignment = align_center
            else:
                c.alignment = align_left_pad

    # ==========================================
    # 2. Tab: 유스케이스 목록 (Usecase List)
    # ==========================================
    ws_uc = wb.create_sheet("유스케이스 목록")
    ws_uc.views.sheetView[0].showGridLines = True
    
    # Column Widths
    ws_uc.column_dimensions['A'].width = 3
    ws_uc.column_dimensions['B'].width = 18  # 유스케이스 ID
    ws_uc.column_dimensions['C'].width = 28  # 유스케이스명
    ws_uc.column_dimensions['D'].width = 20  # 액터
    ws_uc.column_dimensions['E'].width = 32  # 관련 화면/URL
    ws_uc.column_dimensions['F'].width = 48  # 상세설명(개요)
    
    # Title
    ws_uc.row_dimensions[2].height = 36
    ws_uc.merge_cells("B2:F2")
    t_cell = ws_uc["B2"]
    t_cell.value = "2.1 유스케이스 목록 (Usecase List)"
    t_cell.font = title_font
    t_cell.alignment = align_left_pad
    for col in range(2, 7):
        ws_uc.cell(row=2, column=col).border = bottom_heavy_border
        
    # Table Headers
    ws_uc.row_dimensions[4].height = 24
    uc_headers = ["유스케이스 ID", "유스케이스명", "액터", "관련 화면/URL", "상세설명(개요)"]
    for idx, h in enumerate(uc_headers, 2):
        c = ws_uc.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = thin_border
        
    # Usecase Data
    usecases = [
        ("UC-01", "소셜 로그인/회원가입", "비로그인 사용자", "/auth/login, /auth/oauth2", "구글 및 네이버 소셜 계정을 연동한 간편 회원가입 및 로그인을 처리한다."),
        ("UC-02", "프로필 조회 및 수정", "일반 회원", "/mypage/profile", "닉네임 수정 및 연동된 소셜 계정의 기본 정보를 조회한다."),
        ("UC-03", "러닝 기록 등록", "일반 회원", "/running/write", "날짜, 시간, 거리, 페이스, 심박, RPE(강도), 날씨 환경을 수동 또는 API로 등록한다."),
        ("UC-04", "러닝 기록 목록 조회", "일반 회원", "/running/list", "일별 기록 상세 목록 및 주간/월간 누적 거리/횟수/평균 페이스를 통합 조회한다."),
        ("UC-05", "러닝 기록 수정/삭제", "일반 회원", "/running/edit, /running/delete", "자신이 등록한 러닝 기록을 안전하게 수정하고 삭제(소유권 검증)한다."),
        ("UC-06", "월간 목표 설정", "일반 회원", "/goals/monthly", "연/월 기준으로 목표 거리(km) 및 월간 목표 운동 횟수를 관리한다."),
        ("UC-07", "레이스 목표 설정", "일반 회원", "/goals/race", "5k, 10k, 하프, 풀코스 등의 타겟 대회 목표 시간을 설정하고 페이스를 계산한다."),
        ("UC-08", "마라톤 대회 기록 등록", "일반 회원", "/race/register", "대회명, 대회일, 거리, 완주기록 등록 및 공식 기록증 이미지를 업로드한다."),
        ("UC-09", "대회 스케줄 및 D-Day 조회", "일반 회원", "/race/list", "참가 예정인 대회 일정 D-Day 및 완료된 대회 이력을 통합 조회한다."),
        ("UC-10", "개인 최고 기록(PB) 조회", "일반 회원", "/stats/pb", "훈련 및 대회 완주 데이터 중 거리별 개인 최고 기록을 자동 갱신하여 확인한다."),
        ("UC-11", "훈련 데이터 시각화 분석", "일반 회원", "/stats/charts", "거리 추이, 훈련 유형 비율, 온도/날씨별 페이스 상관관계를 차트로 분석한다."),
        ("UC-12", "동물 캐릭터 레벨 조회", "일반 회원", "/stats/level", "누적 주행 거리에 따른 게이미케이션 등급(고양이~치타) 현황을 조회한다.")
    ]
    for r_idx, record in enumerate(usecases, 5):
        ws_uc.row_dimensions[r_idx].height = 20
        row_fill = zebra_fill if r_idx % 2 == 0 else white_fill
        for c_idx, val in enumerate(record, 2):
            c = ws_uc.cell(row=r_idx, column=c_idx, value=val)
            c.font = data_font
            c.fill = row_fill
            c.border = thin_border
            if c_idx in (2, 4):  # ID, Actor
                c.alignment = align_center
            else:
                c.alignment = align_left_pad

    # ==========================================
    # 3. Tab: 기능 요구사항 (FR)
    # ==========================================
    ws_fr = wb.create_sheet("기능 요구사항 (FR)")
    ws_fr.views.sheetView[0].showGridLines = True
    
    # Column Widths
    ws_fr.column_dimensions['A'].width = 3
    ws_fr.column_dimensions['B'].width = 18  # 요구사항 ID
    ws_fr.column_dimensions['C'].width = 28  # 요구사항명
    ws_fr.column_dimensions['D'].width = 12  # 우선순위
    ws_fr.column_dimensions['E'].width = 65  # 상세 설명
    
    # Title
    ws_fr.row_dimensions[2].height = 36
    ws_fr.merge_cells("B2:E2")
    t_cell = ws_fr["B2"]
    t_cell.value = "2.3 기능 요구사항 명세표 (Functional Requirements)"
    t_cell.font = title_font
    t_cell.alignment = align_left_pad
    for col in range(2, 6):
        ws_fr.cell(row=2, column=col).border = bottom_heavy_border
        
    # Table Headers
    ws_fr.row_dimensions[4].height = 24
    fr_headers = ["요구사항 ID", "요구사항명", "우선순위", "상세 설명"]
    for idx, h in enumerate(fr_headers, 2):
        c = ws_fr.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = thin_border
        
    # FR Data
    fr_data = [
        ("FR-01", "소셜 로그인 연동", "필수", "구글 및 네이버 소셜 계정 로그인 연동(OAuth2 프로토콜 규격 사용)"),
        ("FR-02", "러닝 기록 등록 및 기상 연동", "필수", "날짜, 시간, 거리, 페이스, 심박, RPE(운동강도 1~10) 및 기상 API 연동 환경 정보 등록"),
        ("FR-03", "러닝 기록 수정 및 삭제", "필수", "작성자 본인이 등록한 훈련 데이터에 한해 수정 및 삭제 기능 제공 (소유권 유효성 검증)"),
        ("FR-04", "일별/주간/월간 기록 조회", "필수", "기록 리스트 페이징 조회 및 주간/월간 단위 누적 거리 및 평균 페이스 합산 조회"),
        ("FR-05", "월간 목표 및 대회 타겟 설정", "필수", "연/월별 정량적 주행 목표 설정 및 5k/10k/하프/풀코스 타겟 페이스 계획 관리"),
        ("FR-06", "마라톤 대회 및 기록증 보관", "필수", "대회 기본 스케줄 등록 및 증빙용 공식 기록증 이미지 업로드 기능 (최대 10MB)"),
        ("FR-07", "개인 최고 기록(PB) 자동 추출", "필수", "기록 데이터 연산을 통해 거리별 최고 완주 속도(PB) 상시 자동 갱신 및 표출"),
        ("FR-08", "대시보드 통계 그래프 시각화", "보통", "주간, 월간, 연간 누적 거리 추이 선/바 차트 및 전월/전년 동월 대비 분석 통계 시각화"),
        ("FR-09", "날씨 조건별 페이스 영향도 분석", "낮음", "기온, 습도, 날씨 맑음/비 상태에 따른 평균 페이스 영향도 및 비교 통계 분석"),
        ("FR-10", "캐릭터 기반 러닝 레벨 시스템", "낮음", "사용자의 전체 누적 거리에 맞추어 동물 캐릭터 등급(고양이~치타) 자동 부여 및 조회")
    ]
    for r_idx, record in enumerate(fr_data, 5):
        ws_fr.row_dimensions[r_idx].height = 20
        row_fill = zebra_fill if r_idx % 2 == 0 else white_fill
        
        req_id, req_name, priority, desc = record
        
        c_id = ws_fr.cell(row=r_idx, column=2, value=req_id)
        c_name = ws_fr.cell(row=r_idx, column=3, value=req_name)
        c_prio = ws_fr.cell(row=r_idx, column=4, value=priority)
        c_desc = ws_fr.cell(row=r_idx, column=5, value=desc)
        
        # General Styles
        for c in (c_id, c_name, c_prio, c_desc):
            c.font = data_font
            c.fill = row_fill
            c.border = thin_border
            
        c_id.alignment = align_center
        c_name.alignment = align_left_pad
        c_prio.alignment = align_center
        c_desc.alignment = align_left_pad
        
        # Priority Formatting
        if priority == "필수":
            c_prio.fill = fill_essential
            c_prio.font = font_essential
        elif priority == "보통":
            c_prio.fill = fill_normal
            c_prio.font = font_normal
        elif priority == "낮음":
            c_prio.fill = fill_low
            c_prio.font = font_low

    # ==========================================
    # 4. Tab: 비기능 요구사항 (NFR)
    # ==========================================
    ws_nfr = wb.create_sheet("비기능 요구사항 (NFR)")
    ws_nfr.views.sheetView[0].showGridLines = True
    
    # Column Widths
    ws_nfr.column_dimensions['A'].width = 3
    ws_nfr.column_dimensions['B'].width = 18  # 요구사항 ID
    ws_nfr.column_dimensions['C'].width = 20  # 구분
    ws_nfr.column_dimensions['D'].width = 75  # 요구사항 내용
    
    # Title
    ws_nfr.row_dimensions[2].height = 36
    ws_nfr.merge_cells("B2:D2")
    t_cell = ws_nfr["B2"]
    t_cell.value = "2.4 비기능 요구사항 명세표 (Non-Functional Requirements)"
    t_cell.font = title_font
    t_cell.alignment = align_left_pad
    for col in range(2, 5):
        ws_nfr.cell(row=2, column=col).border = bottom_heavy_border
        
    # Table Headers
    ws_nfr.row_dimensions[4].height = 24
    nfr_headers = ["요구사항 ID", "구분", "요구사항 내용"]
    for idx, h in enumerate(nfr_headers, 2):
        c = ws_nfr.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = thin_border
        
    # NFR Data
    nfr_data = [
        ("NFR-01", "보안 (인증)", "OAuth2 소셜 연동을 단일 가입 방식으로 지원하며, Spring Security 설정을 통해 인가 프로세스를 관리한다."),
        ("NFR-02", "보안 (권한)", "로그인하지 않은 외부 방문자의 개인 기록 조회 및 이미지 파일 URI 직접 접근을 서버 단에서 거부(Forbidden) 처리한다."),
        ("NFR-03", "데이터 무결성", "회원이 탈퇴하는 경우, 해당 사용자가 업로드한 기록증 이미지 파일과 모든 훈련 기록 데이터를 연쇄적으로 제거(CASCADE)한다."),
        ("NFR-04", "성능 (대시보드)", "사용자별 훈련 목록 및 대시보드 메인 페이지의 조회 최적화를 위해 데이터베이스 인덱싱을 적용하고, 평균 응답 1.5초를 유지한다."),
        ("NFR-05", "호환성 (데이터베이스)", "MyBatis SQL 매퍼 쿼리 작성 시 표준 ANSI SQL을 준수하여 MySQL 8 및 Oracle DB 다중 데이터베이스 상에서 원활히 동작하도록 보장한다."),
        ("NFR-06", "사용성 (반응형 UI)", "데스크톱 브라우저뿐만 아니라 스마트폰 브라우저 화면 크기에 맞추어 레이아웃이 자동 조정되도록 반응형 그리드를 구현한다."),
        ("NFR-07", "확장성 (가민 워치 연동)", "향후 GPS 가민 스마트 워치 기기에서 추출되는 FIT 파일의 업로드 및 파싱 모듈 연동이 용이하도록 계층적 결합도를 낮추어 설계한다."),
        ("NFR-08", "오류 예외 처리", "서버 비즈니스 오류 발생 시 사용자용 커스텀 응답 구조(Error Response DTO) 및 HTTP Status Code를 일관성 있게 규격화하여 전달한다."),
        ("NFR-09", "확장성 (AI 모델)", "추후 v3의 AI 러닝 코칭(목표 가능성 분석 등) 학습을 대비해 일일 기록 RPE 및 환경 정보의 정형 데이터 규격을 통일하여 보관한다."),
        ("NFR-10", "운영 (로깅)", "SLF4J/Logback 프레임워크를 기반으로 주요 데이터 변경 트랜잭션, 파일 업로드 실패, 인증 예외 상황에 대한 추적 로그를 유지한다.")
    ]
    for r_idx, record in enumerate(nfr_data, 5):
        ws_nfr.row_dimensions[r_idx].height = 20
        row_fill = zebra_fill if r_idx % 2 == 0 else white_fill
        for c_idx, val in enumerate(record, 2):
            c = ws_nfr.cell(row=r_idx, column=c_idx, value=val)
            c.font = data_font
            c.fill = row_fill
            c.border = thin_border
            if c_idx in (2, 3):  # ID, Category
                c.alignment = align_center
                if c_idx == 3:
                    c.font = Font(name=font_name, size=9.5, bold=True, color="1A365D")
            else:
                c.alignment = align_left_pad

    filename = "요구사항_명세서_Running_Coach_v1.1.xlsx"
    wb.save(filename)
    print(f"Successfully created revised Excel Document '{filename}'")

if __name__ == "__main__":
    main()
