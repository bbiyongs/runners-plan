import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    wb = openpyxl.Workbook()
    
    # Remove default sheet to build custom order
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Shared Style Declarations
    font_name = "Malgun Gothic"
    title_font = Font(name=font_name, size=14, bold=True, color="1A365D")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    category_font = Font(name=font_name, size=10, bold=True, color="1A365D")
    data_font = Font(name=font_name, size=9.5, color="2D3748")
    
    # Fills
    fills = {
        'header': PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid"),
        'category': PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid"),
        'white': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        'zebra': PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"),
        'auth_y': PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"), # Light Red
        'auth_n': PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")  # Light Gray
    }
    
    # Fonts for specific fields
    auth_y_font = Font(name=font_name, size=9.5, bold=True, color="9B2C2C") # Dark Red
    auth_n_font = Font(name=font_name, size=9.5, color="4A5568") # Dark Gray
    
    # Method Styles (Swagger-like styling)
    method_styles = {
        'GET': {
            'font': Font(name=font_name, size=9.5, bold=True, color="2B6CB0"), # Blue
            'fill': PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
        },
        'POST': {
            'font': Font(name=font_name, size=9.5, bold=True, color="228B22"), # Forest Green
            'fill': PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        },
        'PUT': {
            'font': Font(name=font_name, size=9.5, bold=True, color="B7791F"), # Dark Golden
            'fill': PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")
        },
        'DELETE': {
            'font': Font(name=font_name, size=9.5, bold=True, color="C53030"), # Red
            'fill': PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
        }
    }

    alignments = {
        'center': Alignment(horizontal="center", vertical="center"),
        'left_pad': Alignment(horizontal="left", vertical="center", indent=1),
        'left_wrap': Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    }
    
    borders = {
        'thin': Border(left=Side(style='thin', color='CBD5E0'), right=Side(style='thin', color='CBD5E0'),
                       top=Side(style='thin', color='CBD5E0'), bottom=Side(style='thin', color='CBD5E0')),
        'heavy': Border(bottom=Side(style='medium', color='1A365D'))
    }

    # ==========================================
    # 1. Tab: 버전관리 (Version History)
    # ==========================================
    ws_ver = wb.create_sheet("버전관리")
    ws_ver.views.sheetView[0].showGridLines = True
    ws_ver.column_dimensions['A'].width = 3
    ws_ver.column_dimensions['B'].width = 12
    ws_ver.column_dimensions['C'].width = 16
    ws_ver.column_dimensions['D'].width = 48
    ws_ver.column_dimensions['E'].width = 22
    ws_ver.column_dimensions['F'].width = 15
    
    ws_ver.row_dimensions[2].height = 36
    ws_ver.merge_cells("B2:F2")
    t_cell = ws_ver["B2"]
    t_cell.value = "API 목록 명세서 - 버전 관리 이력"
    t_cell.font = title_font
    t_cell.alignment = alignments['left_pad']
    for col in range(2, 7):
        ws_ver.cell(row=2, column=col).border = borders['heavy']
        
    ws_ver.row_dimensions[4].height = 24
    for idx, h in enumerate(["버전", "변경 일자", "변경 내용", "작성자", "비고"], 2):
        c = ws_ver.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    ver_records = [
        ("v1.0", "2026-06-24", "프로젝트 상세 설계안 기준 최초 API 목록 작성", "Running Coach 개발팀", "-")
    ]
    for r_idx, record in enumerate(ver_records, 5):
        ws_ver.row_dimensions[r_idx].height = 20
        row_fill = fills['zebra'] if r_idx % 2 == 0 else fills['white']
        for c_idx, val in enumerate(record, 2):
            c = ws_ver.cell(row=r_idx, column=c_idx, value=val)
            c.font = data_font
            c.fill = row_fill
            c.border = borders['thin']
            if c_idx in (2, 3, 5):
                c.alignment = alignments['center']
            else:
                c.alignment = alignments['left_pad']

    # ==========================================
    # 2. Tab: API 목록 (API List)
    # ==========================================
    ws_api = wb.create_sheet("API 목록")
    ws_api.views.sheetView[0].showGridLines = True
    ws_api.column_dimensions['A'].width = 3
    ws_api.column_dimensions['B'].width = 16  # No
    ws_api.column_dimensions['C'].width = 24  # 기능명
    ws_api.column_dimensions['D'].width = 14  # Method
    ws_api.column_dimensions['E'].width = 38  # URL
    ws_api.column_dimensions['F'].width = 12  # 인증여부
    ws_api.column_dimensions['G'].width = 48  # 설명
    
    # Sheet Title
    ws_api.row_dimensions[2].height = 36
    ws_api.merge_cells("B2:G2")
    title_cell = ws_api["B2"]
    title_cell.value = "프로젝트 백엔드 API 목록 명세서 (v1.0)"
    title_cell.font = title_font
    title_cell.alignment = alignments['left_pad']
    for col in range(2, 8):
        ws_api.cell(row=2, column=col).border = borders['heavy']
        
    # Table Headers
    ws_api.row_dimensions[4].height = 24
    api_headers = ["No", "기능명", "Method", "URL", "인증", "설명"]
    for idx, h in enumerate(api_headers, 2):
        c = ws_api.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    # API Data structure by category
    api_data = [
        # Category 1
        {
            "category": "1. 인증 (Auth)",
            "items": [
                ("AUTH-001", "Google 로그인", "POST", "/api/auth/google", "N", "Google OAuth 로그인 처리"),
                ("AUTH-002", "Naver 로그인", "POST", "/api/auth/naver", "N", "Naver OAuth 로그인 처리"),
                ("AUTH-003", "로그아웃", "POST", "/api/auth/logout", "Y", "로그인 세션 만료 및 로그아웃"),
                ("AUTH-004", "내 정보 조회", "GET", "/api/auth/me", "Y", "로그인된 사용자의 프로필 정보 조회"),
                ("AUTH-005", "회원정보 수정", "PUT", "/api/auth/me", "Y", "로그인 사용자의 닉네임 수정")
            ]
        },
        # Category 2
        {
            "category": "2. 러닝 기록 (Run Record)",
            "items": [
                ("RUN-001", "러닝 기록 등록", "POST", "/api/run-records", "Y", "신규 러닝 훈련 기록 등록 및 날씨 정보 자동 매핑"),
                ("RUN-002", "러닝 기록 수정", "PUT", "/api/run-records/{runRecordId}", "Y", "기존 러닝 기록 정보 수정 (본인 소유 기록 검증)"),
                ("RUN-003", "러닝 기록 삭제", "DELETE", "/api/run-records/{runRecordId}", "Y", "러닝 기록 삭제 (본인 소유 기록 검증)"),
                ("RUN-004", "러닝 기록 상세조회", "GET", "/api/run-records/{runRecordId}", "Y", "특정 러닝 기록의 상세 데이터 단건 조회"),
                ("RUN-005", "일별 기록 조회", "GET", "/api/run-records/daily", "Y", "특정 날짜 범위 내의 러닝 기록 조회"),
                ("RUN-006", "주간 기록 조회", "GET", "/api/run-records/weekly", "Y", "최근 주간 주행 거리 및 평균 수치 집계 조회"),
                ("RUN-007", "월간 기록 조회", "GET", "/api/run-records/monthly", "Y", "월간 누적 거리, 평균 페이스 및 심박수 집계 조회"),
                ("RUN-008", "최근 기록 조회", "GET", "/api/run-records/recent", "Y", "메인 대시보드 표출용 최근 운동 기록 목록 조회 (최신순)")
            ]
        },
        # Category 3
        {
            "category": "3. 목표 관리 (Goal)",
            "items": [
                ("GOAL-001", "목표 등록", "POST", "/api/goals", "Y", "신규 월간 또는 레이스 완주 목표 설정 등록"),
                ("GOAL-002", "목표 수정", "PUT", "/api/goals/{goalId}", "Y", "기존 설정 목표 내용 수정"),
                ("GOAL-003", "목표 삭제", "DELETE", "/api/goals/{goalId}", "Y", "설정 목표 레코드 삭제"),
                ("GOAL-004", "목표 상세조회", "GET", "/api/goals/{goalId}", "Y", "단일 목표 상세 정보 조회"),
                ("GOAL-005", "목표 목록 조회", "GET", "/api/goals", "Y", "설정 목표 전체 목록 조회"),
                ("GOAL-006", "목표 달성률 조회", "GET", "/api/goals/progress", "Y", "현재 기준 월간 목표 달성률 계산 및 조회")
            ]
        },
        # Category 4
        {
            "category": "4. 대회 기록 (Race Record)",
            "items": [
                ("RACE-001", "대회 기록 등록", "POST", "/api/race-records", "Y", "마라톤 참가 대회 기록 및 공식 기록증 파일 경로 등록"),
                ("RACE-002", "대회 기록 수정", "PUT", "/api/race-records/{raceRecordId}", "Y", "마라톤 대회 기록 데이터 및 파일 수정"),
                ("RACE-003", "대회 기록 삭제", "DELETE", "/api/race-records/{raceRecordId}", "Y", "대회 기록 레코드 삭제"),
                ("RACE-004", "대회 기록 상세조회", "GET", "/api/race-records/{raceRecordId}", "Y", "특정 대회 기록 단건 상세 조회"),
                ("RACE-005", "대회 기록 목록조회", "GET", "/api/race-records", "Y", "참가한 마라톤 대회 기록 전체 목록 조회"),
                ("RACE-006", "PB 조회", "GET", "/api/race-records/pb", "Y", "5k, 10k, 하프, 풀코스 거리별 개인 최고 기록 조회")
            ]
        },
        # Category 5
        {
            "category": "5. 통계 (Statistics)",
            "items": [
                ("STAT-001", "대시보드 조회", "GET", "/api/statistics/dashboard", "Y", "대시보드 화면용 요약 집계 데이터 및 PB 현황 일괄 조회"),
                ("STAT-002", "월별 분석 조회", "GET", "/api/statistics/monthly", "Y", "월별 주행 거리, 시간, 횟수 통계 조회"),
                ("STAT-003", "전월 비교 조회", "GET", "/api/statistics/monthly/compare", "Y", "전월 대비 주행 거리 및 페이스 변화율 비교"),
                ("STAT-004", "전년도 비교 조회", "GET", "/api/statistics/yearly/compare", "Y", "전년 동월 대비 데이터 비교 조회"),
                ("STAT-005", "페이스 추이 조회", "GET", "/api/statistics/pace-trend", "Y", "일별/주간 페이스 변화 흐름 분석 데이터 조회"),
                ("STAT-006", "거리 추이 조회", "GET", "/api/statistics/distance-trend", "Y", "주간/월간 주행 거리 증감 추이 분석 데이터 조회"),
                ("STAT-007", "훈련유형 분석", "GET", "/api/statistics/training-type", "Y", "이지런, LSD, 인터벌 등 훈련 스타일 비중 백분율 조회"),
                ("STAT-008", "러닝 레벨 분석", "GET", "/api/statistics/level", "Y", "총 누적 거리에 해당하는 동물 등급 및 레벨업 잔여량 계산")
            ]
        },
        # Category 6
        {
            "category": "6. 공통 코드 (Code)",
            "items": [
                ("CODE-001", "코드그룹 조회", "GET", "/api/codes/groups", "Y", "전체 공통 코드 그룹 목록 조회"),
                ("CODE-002", "상세코드 조회", "GET", "/api/codes/{groupCode}", "Y", "특정 그룹코드에 매핑된 하위 상세 코드 전체 목록 조회")
            ]
        }
    ]
    
    current_row = 5
    for cat_data in api_data:
        # Category Separator Row (Merged B:G)
        ws_api.row_dimensions[current_row].height = 24
        ws_api.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
        cat_cell = ws_api.cell(row=current_row, column=2, value=f"  ■ {cat_data['category']}")
        cat_cell.font = category_font
        cat_cell.fill = fills['category']
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Border for merged category cells
        for col in range(2, 8):
            ws_api.cell(row=current_row, column=col).border = borders['thin']
            
        current_row += 1
        
        # Data Rows
        for item_idx, item in enumerate(cat_data['items']):
            ws_api.row_dimensions[current_row].height = 20
            row_fill = fills['zebra'] if item_idx % 2 == 1 else fills['white']
            no, func_name, method, url, auth_req, desc = item
            
            # 1. Write values
            ws_api.cell(row=current_row, column=2, value=no)
            ws_api.cell(row=current_row, column=3, value=func_name)
            ws_api.cell(row=current_row, column=4, value=method)
            ws_api.cell(row=current_row, column=5, value=url)
            ws_api.cell(row=current_row, column=6, value=auth_req)
            ws_api.cell(row=current_row, column=7, value=desc)
            
            # 2. Format cells
            for col in range(2, 8):
                c = ws_api.cell(row=current_row, column=col)
                c.font = data_font
                c.fill = row_fill
                c.border = borders['thin']
                
                # Alignments
                if col in (2, 4, 6): # No, Method, Auth
                    c.alignment = alignments['center']
                else: # Name, URL, Description
                    c.alignment = alignments['left_pad']
            
            # 3. Apply custom HTTP Method Colors
            if method in method_styles:
                m_cell = ws_api.cell(row=current_row, column=4)
                m_cell.font = method_styles[method]['font']
                m_cell.fill = method_styles[method]['fill']
                
            # 4. Apply custom Auth styles
            auth_cell = ws_api.cell(row=current_row, column=6)
            if auth_req == 'Y':
                auth_cell.font = auth_y_font
                auth_cell.fill = fills['auth_y']
            else:
                auth_cell.font = auth_n_font
                auth_cell.fill = fills['auth_n']
                
            current_row += 1
            
        # Spacer row between sections
        current_row += 1
        
    filename = "C:/Runners_plan/산출물/API_목록_Running_Coach_v1.0.xlsx"
    wb.save(filename)
    print(f"Successfully created API list Excel sheet '{filename}'")

if __name__ == "__main__":
    main()
