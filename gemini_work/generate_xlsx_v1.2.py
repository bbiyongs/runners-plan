import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def add_usecase_detail_table(ws, start_row, uc_id, uc_name, actor, precondition, basic_flow, exception_flow, postcondition, fonts, fills, borders, alignments):
    # Title Row (merged B{start_row}:C{start_row})
    ws.row_dimensions[start_row].height = 28
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)
    title_cell = ws.cell(row=start_row, column=2, value=f"  ■ {uc_id}. {uc_name}")
    title_cell.font = fonts['title']
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Bottom heavy border for title row B and C
    ws.cell(row=start_row, column=2).border = Border(bottom=Side(style='medium', color='1A365D'))
    ws.cell(row=start_row, column=3).border = Border(bottom=Side(style='medium', color='1A365D'))
    
    # Header Row (start_row + 1)
    header_row = start_row + 1
    ws.row_dimensions[header_row].height = 22
    h1 = ws.cell(row=header_row, column=2, value="항목")
    h2 = ws.cell(row=header_row, column=3, value="내용")
    for h in (h1, h2):
        h.font = fonts['header']
        h.fill = fills['header']
        h.alignment = alignments['center']
        h.border = borders['thin']
        
    items = [
        ("액터", actor),
        ("사전조건", precondition),
        ("기본 흐름", basic_flow),
        ("예외/대안 흐름", exception_flow),
        ("사후조건", postcondition)
    ]
    
    curr_row = header_row + 1
    for label, val in items:
        lines = val.count('\n') + 1
        ws.row_dimensions[curr_row].height = max(22, lines * 17)
        
        c_lbl = ws.cell(row=curr_row, column=2, value=label)
        c_val = ws.cell(row=curr_row, column=3, value=val)
        
        c_lbl.font = fonts['label']
        c_lbl.fill = fills['label']
        c_lbl.alignment = alignments['center']
        c_lbl.border = borders['thin']
        
        c_val.font = fonts['data']
        c_val.fill = fills['white']
        c_val.alignment = alignments['left_wrap']
        c_val.border = borders['thin']
        
        curr_row += 1
        
    return curr_row + 1 # return the next start row (leaving 1 empty row)

def main():
    wb = openpyxl.Workbook()
    
    # Remove default sheet to build custom order
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Shared Style Declarations
    font_name = "Malgun Gothic"
    fonts = {
        'sheet_title': Font(name=font_name, size=14, bold=True, color="1A365D"),
        'title': Font(name=font_name, size=11, bold=True, color="1A365D"),
        'header': Font(name=font_name, size=10, bold=True, color="FFFFFF"),
        'label': Font(name=font_name, size=9.5, bold=True, color="1A365D"),
        'data': Font(name=font_name, size=9.5, color="2D3748")
    }
    fills = {
        'header': PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid"),
        'label': PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid"),
        'white': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        'zebra': PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"),
        'section': PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    }
    alignments = {
        'center': Alignment(horizontal="center", vertical="center"),
        'left_pad': Alignment(horizontal="left", vertical="center", indent=1),
        'left_wrap': Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    }
    borders = {
        'thin': Border(left=Side(style='thin', color='CBD5E0'), right=Side(style='thin', color='CBD5E0'),
                       top=Side(style='thin', color='CBD5E0'), bottom=Side(style='thin', color='CBD5E0')),
        'heavy': Border(bottom=Side(style='medium', color='1A365D'))
    }

    # Priority colors (soft background with matching dark text)
    fill_essential = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid") # 필수 (soft red)
    font_essential = Font(name=font_name, size=9.5, bold=True, color="9B2C2C")
    
    fill_normal = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid") # 보통 (soft yellow)
    font_normal = Font(name=font_name, size=9.5, bold=True, color="975A16")
    
    fill_low = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # 낮음 (soft gray)
    font_low = Font(name=font_name, size=9.5, bold=True, color="4A5568")

    # ==========================================
    # 1. Tab: 버전관리 (Version History)
    # ==========================================
    ws_ver = wb.create_sheet("버전관리")
    ws_ver.views.sheetView[0].showGridLines = True
    ws_ver.column_dimensions['A'].width = 3
    ws_ver.column_dimensions['B'].width = 12
    ws_ver.column_dimensions['C'].width = 16
    ws_ver.column_dimensions['D'].width = 48
    ws_ver.column_dimensions['E'].width = 22
    ws_ver.column_dimensions['F'].width = 15
    
    ws_ver.row_dimensions[2].height = 36
    ws_ver.merge_cells("B2:F2")
    t_cell = ws_ver["B2"]
    t_cell.value = "요구사항 명세서 - 버전 관리 이력"
    t_cell.font = fonts['sheet_title']
    t_cell.alignment = alignments['left_pad']
    for col in range(2, 7):
        ws_ver.cell(row=2, column=col).border = borders['heavy']
        
    ws_ver.row_dimensions[4].height = 24
    for idx, h in enumerate(["버전", "변경 일자", "변경 내용", "작성자", "비고"], 2):
        c = ws_ver.cell(row=4, column=idx, value=h)
        c.font = fonts['header']
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    ver_records = [
        ("v1.0", "2026-06-18", "최초 요구사항 정의서 작성 및 멀티 탭 문서화", "Running Coach 개발팀", "-"),
        ("v1.1", "2026-06-22", "산출물 가이드(5-10p) 요구 사양 기준 개정 (Usecase, FR, NFR 분리)", "Running Coach 기획팀", "-"),
        ("v1.2", "2026-06-23", "기능별 유스케이스 상세 명세 탭(5종) 신설 및 데이터 작성", "Running Coach 기획팀", "-")
    ]
    for r_idx, record in enumerate(ver_records, 5):
        ws_ver.row_dimensions[r_idx].height = 20
        row_fill = fills['zebra'] if r_idx % 2 == 0 else fills['white']
        for c_idx, val in enumerate(record, 2):
            c = ws_ver.cell(row=r_idx, column=c_idx, value=val)
            c.font = fonts['data']
            c.fill = row_fill
            c.border = borders['thin']
            if c_idx in (2, 3, 5):
                c.alignment = alignments['center']
            else:
                c.alignment = alignments['left_pad']

    # ==========================================
    # 2. Tab: 유스케이스 목록 (Usecase List)
    # ==========================================
    ws_uc = wb.create_sheet("유스케이스 목록")
    ws_uc.views.sheetView[0].showGridLines = True
    ws_uc.column_dimensions['A'].width = 3
    ws_uc.column_dimensions['B'].width = 18
    ws_uc.column_dimensions['C'].width = 28
    ws_uc.column_dimensions['D'].width = 20
    ws_uc.column_dimensions['E'].width = 32
    ws_uc.column_dimensions['F'].width = 48
    
    ws_uc.row_dimensions[2].height = 36
    ws_uc.merge_cells("B2:F2")
    t_cell = ws_uc["B2"]
    t_cell.value = "2.1 유스케이스 목록 (Usecase List)"
    t_cell.font = fonts['sheet_title']
    t_cell.alignment = alignments['left_pad']
    for col in range(2, 7):
        ws_uc.cell(row=2, column=col).border = borders['heavy']
        
    ws_uc.row_dimensions[4].height = 24
    for idx, h in enumerate(["유스케이스 ID", "유스케이스명", "액터", "관련 화면/URL", "상세설명(개요)"], 2):
        c = ws_uc.cell(row=4, column=idx, value=h)
        c.font = fonts['header']
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    usecases = [
        ("UC-01", "소셜 로그인/회원가입", "비로그인 사용자", "/auth/login, /auth/oauth2", "구글 및 네이버 소셜 계정을 연동한 간편 회원가입 및 로그인을 처리한다."),
        ("UC-02", "프로필 조회 및 수정", "일반 회원", "/mypage/profile", "닉네임 수정 및 연동된 소셜 계정의 기본 정보를 조회한다."),
        ("UC-03", "러닝 기록 등록", "일반 회원", "/running/write", "날짜, 시간, 거리, 페이스, 심박, RPE(강도), 날씨 환경을 수동 또는 API로 등록한다."),
        ("UC-04", "러닝 기록 목록 조회", "일반 회원", "/running/list", "일별 기록 상세 목록 및 주간/월간 누적 거리/횟수/평균 페이스를 통합 조회한다."),
        ("UC-05", "러닝 기록 수정/삭제", "일반 회원", "/running/edit, /running/delete", "자신이 등록한 러닝 기록을 안전하게 수정하고 삭제(소유권 검증)한다."),
        ("UC-06", "월간 목표 설정", "일반 회원", "/goals/monthly", "연/월 기준으로 목표 거리(km) 및 월간 목표 운동 횟수를 관리한다."),
        ("UC-07", "레이스 목표 설정", "일반 회원", "/goals/race", "5k, 10k, 하프, 풀코스 등의 타겟 대회 목표 시간을 설정하고 페이스를 계산한다."),
        ("UC-08", "마라톤 대회 기록 등록", "일반 회원", "/race/register", "대회명, 대회일, 거리, 완주기록 등록 및 공식 기록증 이미지를 업로드한다."),
        ("UC-09", "대회 스케줄 및 D-Day 조회", "일반 회원", "/race/list", "참가 예정인 대회 일정 D-Day 및 완료된 대회 이력을 통합 조회한다."),
        ("UC-10", "개인 최고 기록(PB) 조회", "일반 회원", "/stats/pb", "훈련 및 대회 완주 데이터 중 거리별 개인 최고 기록을 자동 갱신하여 확인한다."),
        ("UC-11", "훈련 데이터 시각화 분석", "일반 회원", "/stats/charts", "거리 추이, 훈련 유형 비율, 온도/날씨별 페이스 상관관계를 차트로 분석한다."),
        ("UC-12", "동물 캐릭터 레벨 조회", "일반 회원", "/stats/level", "누적 주행 거리에 따른 게이미케이션 등급(고양이~치타) 현황을 조회한다.")
    ]
    for r_idx, record in enumerate(usecases, 5):
        ws_uc.row_dimensions[r_idx].height = 20
        row_fill = fills['zebra'] if r_idx % 2 == 0 else fills['white']
        for c_idx, val in enumerate(record, 2):
            c = ws_uc.cell(row=r_idx, column=c_idx, value=val)
            c.font = fonts['data']
            c.fill = row_fill
            c.border = borders['thin']
            if c_idx in (2, 4):
                c.alignment = alignments['center']
            else:
                c.alignment = alignments['left_pad']

    # ==========================================
    # 3. Tabs: Usecase Detailed Specifications
    # ==========================================
    
    # A. UC_회원 Tab
    ws_uch = wb.create_sheet("UC_회원")
    ws_uch.views.sheetView[0].showGridLines = True
    ws_uch.column_dimensions['A'].width = 3
    ws_uch.column_dimensions['B'].width = 18
    ws_uch.column_dimensions['C'].width = 80
    
    row = add_usecase_detail_table(
        ws_uch, 2, "UC-01", "소셜 로그인/회원가입",
        "비로그인 사용자",
        "네이버 또는 구글 계정을 보유하고 있어야 한다.",
        "1) 사용자가 로그인 페이지에서 소셜 로그인 버튼(구글/네이버)을 클릭한다.\n"
        "2) 해당 소셜 인증 서버로 리다이렉트 되어 사용자가 인증 절차를 수행한다.\n"
        "3) 소셜 로그인 성공 시 우리 서비스 서버로 Authorization Code가 전달되며, 백엔드에서 사용자 프로필을 조회한다.\n"
        "4) 처음 로그인한 경우 자동으로 회원 DB에 회원가입 등록(USER 테이블)을 완료한다.\n"
        "5) 가입 완료 또는 기존 사용자 로그인 완료 후 JWT 토큰 또는 세션을 발행하고 대시보드로 이동한다.",
        "2a) 소셜 인증 과정에서 사용자가 인증을 거부하거나 오류가 발생한 경우:\n"
        "    - 오류 메시지를 출력하고 로그인 메인 화면으로 돌아간다.",
        "로그인 세션/토큰이 생성되며, 회원 권한(ROLE_USER)을 획득한다.",
        fonts, fills, borders, alignments
    )
    
    add_usecase_detail_table(
        ws_uch, row, "UC-02", "프로필 조회 및 수정",
        "일반 회원",
        "소셜 로그인이 완료되어 유효한 세션/토큰을 보유하고 있어야 한다.",
        "1) 사용자가 마이페이지 프로필 관리 탭에 진입한다.\n"
        "2) 시스템이 사용자 ID에 해당하는 닉네임, 소셜 로그인 제공자 및 연동 이메일 정보를 조회하여 화면에 출력한다.\n"
        "3) 사용자가 수정할 닉네임을 입력하고 '저장' 버튼을 클릭한다.\n"
        "4) 시스템이 입력값 유효성을 검사한 후 RUNNER 테이블의 정보를 수정한다.\n"
        "5) 수정이 완료되면 완료 알림을 띄우고 변경된 닉네임으로 화면을 갱신한다.",
        "3a) 중복되거나 유효하지 않은 형식의 닉네임인 경우:\n"
        "    - '유효하지 않은 닉네임입니다' 에러 메시지를 표시하고 입력 대기 상태로 유지한다.",
        "데이터베이스의 사용자 정보가 갱신되며, 변경 사항이 화면에 즉시 표시된다.",
        fonts, fills, borders, alignments
    )

    # B. UC_러닝기록 Tab
    ws_ucr = wb.create_sheet("UC_러닝기록")
    ws_ucr.views.sheetView[0].showGridLines = True
    ws_ucr.column_dimensions['A'].width = 3
    ws_ucr.column_dimensions['B'].width = 18
    ws_ucr.column_dimensions['C'].width = 80
    
    row = add_usecase_detail_table(
        ws_ucr, 2, "UC-03", "러닝 기록 등록",
        "일반 회원",
        "사용자가 로그인 상태여야 한다.",
        "1) 사용자가 러닝 기록 등록 버튼을 클릭하여 입력 폼을 활성화한다.\n"
        "2) 날짜, 거리(km), 운동시간(분:초), 평균 심박수, 훈련종류(이지런, LSD 등), 훈련강도(RPE 1~10), 메모를 입력한다.\n"
        "3) 시스템이 날짜와 시간을 기반으로 기상 API를 호출하여 당시의 온도, 습도, 날씨 상태를 자동으로 수집해 폼에 채운다.\n"
        "4) 사용자가 '등록' 버튼을 클릭하면 RUN_RECORD 테이블에 기록을 저장한다.\n"
        "5) 기록 저장 완료 메시지와 함께 목록 화면으로 리다이렉트 된다.",
        "2a) 필수 입력값(거리, 시간 등) 누락 시:\n"
        "    - '필수 항목을 입력해주세요' 알림과 함께 미입력 필드를 강조한다.\n"
        "3a) 기상 API 장애 또는 일시적인 호출 실패 시:\n"
        "    - 날씨 정보는 공백으로 처리하고 사용자 등록이 정상 차단되지 않도록 기본 기록만 저장한다.",
        "신규 러닝 기록이 데이터베이스(RUN_RECORD 테이블)에 정상 저장된다.",
        fonts, fills, borders, alignments
    )
    
    row = add_usecase_detail_table(
        ws_ucr, row, "UC-04", "러닝 기록 목록 조회",
        "일반 회원",
        "사용자가 로그인 상태여야 한다.",
        "1) 사용자가 러닝 기록 목록 메뉴를 선택한다.\n"
        "2) 시스템이 해당 사용자의 전체 러닝 기록 데이터를 조회한다.\n"
        "3) 기본 일별 기록 목록(날짜, 거리, 페이스, 훈련유형)을 최신순으로 페이징 처리하여 리스트로 출력한다.\n"
        "4) 상단 영역에 주간/월간 단위의 누적 주행 거리와 횟수, 평균 페이스 집계 수치를 시각화하여 표현한다.",
        "2a) 조회된 기록이 없는 경우:\n"
        "    - '등록된 러닝 기록이 없습니다' 안내 메시지를 화면에 출력한다.",
        "해당 사용자의 전체 러닝 내역 및 요약 통계 정보가 화면에 출력된다.",
        fonts, fills, borders, alignments
    )
    
    add_usecase_detail_table(
        ws_ucr, row, "UC-05", "러닝 기록 수정/삭제",
        "일반 회원",
        "사용자가 로그인 상태이고, 수정/삭제하려는 러닝 기록의 소유자여야 한다.",
        "1) 사용자가 특정 러닝 기록 상세 보기 화면에서 '수정' 또는 '삭제' 버튼을 클릭한다.\n"
        "2) 시스템은 현재 로그인한 회원 ID와 해당 레코드의 USER_ID가 일치하는지 소유권을 검증한다.\n"
        "3) [수정 시] 수정 폼을 제공하고, 사용자가 값을 변경한 후 저장하면 데이터베이스를 업데이트한다.\n"
        "4) [삭제 시] 삭제 재확인 팝업을 거쳐 승인하면 데이터베이스에서 해당 레코드를 삭제(Delete)한다.\n"
        "5) 처리 완료 후 목록 화면으로 이동한다.",
        "2a) 다른 사용자의 기록에 접근하거나 수정/삭제를 시도할 경우:\n"
        "    - '권한이 없습니다' 에러 메시지를 반환하고 메인 화면으로 강제 이동시킨다.",
        "데이터베이스(RUN_RECORD 테이블)의 기존 데이터가 갱신되거나 물리적으로 삭제된다.",
        fonts, fills, borders, alignments
    )

    # C. UC_목표관리 Tab
    ws_ucg = wb.create_sheet("UC_목표관리")
    ws_ucg.views.sheetView[0].showGridLines = True
    ws_ucg.column_dimensions['A'].width = 3
    ws_ucg.column_dimensions['B'].width = 18
    ws_ucg.column_dimensions['C'].width = 80
    
    row = add_usecase_detail_table(
        ws_ucg, 2, "UC-06", "월간 목표 설정",
        "일반 회원",
        "사용자가 로그인 상태여야 한다.",
        "1) 사용자가 목표 설정 메뉴에서 '월간 목표 설정' 탭을 선택한다.\n"
        "2) 연/월(예: 2026년 7월), 목표 거리(예: 200km), 목표 운동 횟수(예: 20회)를 설정한다.\n"
        "3) '저장' 버튼을 클릭한다.\n"
        "4) 시스템이 유효성을 검사한 후 GOAL 테이블에 월간 목표 정보를 저장한다.",
        "2a) 이미 동일한 연/월에 목표가 존재하는 경우:\n"
        "    - '기존 목표가 존재합니다. 수정하시겠습니까?' 팝업을 띄우고, 승인 시 수정(Update) 처리한다.",
        "데이터베이스(GOAL 테이블)에 월간 목표 데이터가 성공적으로 반영된다.",
        fonts, fills, borders, alignments
    )
    
    add_usecase_detail_table(
        ws_ucg, row, "UC-07", "레이스 목표 설정",
        "일반 회원",
        "사용자가 로그인 상태여야 한다.",
        "1) 사용자가 목표 설정 메뉴에서 '레이스 목표 설정' 탭을 선택한다.\n"
        "2) 5km, 10km, 하프, 풀코스 중 원하는 대상을 선택한다.\n"
        "3) 목표 시간(예: 10km - 50분)을 입력하고 저장한다.\n"
        "4) 시스템이 입력 시간을 페이스(min/km)로 자동 계산하여 결과를 출력하고, GOAL 테이블의 레이스 필드를 갱신한다.",
        "3a) 음수 또는 불합리한 시간 데이터 입력 시:\n"
        "    - '올바른 완주 시간 형식을 입력해 주세요' 경고를 표시한다.",
        "목표 시간과 자동 계산된 목표 페이스가 저장된다.",
        fonts, fills, borders, alignments
    )

    # D. UC_대회기록 Tab
    ws_ucd = wb.create_sheet("UC_대회기록")
    ws_ucd.views.sheetView[0].showGridLines = True
    ws_ucd.column_dimensions['A'].width = 3
    ws_ucd.column_dimensions['B'].width = 18
    ws_ucd.column_dimensions['C'].width = 80
    
    row = add_usecase_detail_table(
        ws_ucd, 2, "UC-08", "마라톤 대회 기록 등록",
        "일반 회원",
        "사용자가 로그인 상태여야 한다.",
        "1) 사용자가 대회 기록 관리 탭에서 '대회 등록' 버튼을 클릭한다.\n"
        "2) 대회명, 대회 일시, 참가 거리, 완주 기록(시간), 대회 장소 및 공식 홈페이지 URL을 입력한다.\n"
        "3) 증빙용 공식 기록증 이미지 파일을 업로드한다.\n"
        "4) 시스템이 이미지 확장자 및 크기(10MB 이하) 유효성을 검증하고 파일을 업로드한다.\n"
        "5) '저장'을 클릭하면 RACE_RECORD 테이블에 저장 완료 후 목록으로 리다이렉트 된다.",
        "4a) 이미지 파일 용량이 10MB를 초과하거나 허용되지 않은 확장자일 경우:\n"
        "    - '업로드할 수 없는 파일 포맷이거나 파일 크기 초과입니다' 경고를 반환한다.",
        "대회 기록 및 기록증 이미지 파일 경로가 데이터베이스(RACE_RECORD 테이블)에 매핑된다.",
        fonts, fills, borders, alignments
    )
    
    add_usecase_detail_table(
        ws_ucd, row, "UC-09", "대회 스케줄 및 D-Day 조회",
        "일반 회원",
        "사용자가 로그인 상태여야 한다.",
        "1) 사용자가 대회 기록 메뉴로 진입한다.\n"
        "2) 시스템은 사용자가 등록한 완료 대회 기록 목록과 다가오는 예정 대회 일정을 분리하여 조회한다.\n"
        "3) 다가오는 대회의 경우 D-Day(디데이) 카운트를 실시간 계산하여 대시보드 및 리스트 상단에 노출한다.",
        "2a) 예정 또는 완료된 대회 데이터가 없는 경우:\n"
        "    - '참가 신청 예정이거나 완료한 대회 일정이 없습니다' 문구를 출력한다.",
        "사용자의 마라톤 대회 타임라인과 예정 대회의 D-Day 정보가 표기된다.",
        fonts, fills, borders, alignments
    )

    # E. UC_통계레벨 Tab
    ws_ucs = wb.create_sheet("UC_통계레벨")
    ws_ucs.views.sheetView[0].showGridLines = True
    ws_ucs.column_dimensions['A'].width = 3
    ws_ucs.column_dimensions['B'].width = 18
    ws_ucs.column_dimensions['C'].width = 80
    
    row = add_usecase_detail_table(
        ws_ucs, 2, "UC-10", "개인 최고 기록(PB) 조회",
        "일반 회원",
        "사용자가 로그인 상태여야 한다.",
        "1) 사용자가 대시보드 또는 통계 메뉴에 접근한다.\n"
        "2) 시스템은 사용자의 전체 RUN_RECORD 및 RACE_RECORD 데이터를 5km, 10km, 하프, 풀코스 구간으로 필터링한다.\n"
        "3) 각 구간에서 완주 시간이 가장 짧은 레코드를 탐색하여 개인 최고 기록(Personal Best)으로 자동 산출한다.\n"
        "4) 탐색된 각 거리별 PB 시간 및 페이스를 화면에 가독성 있게 노출한다.",
        "2a) 특정 거리에 해당하는 러닝 기록이 전무할 경우:\n"
        "    - 최고 기록 필드에 '기록 없음' 또는 '-' 기호로 표시한다.",
        "거리별 실시간 PB 데이터가 대시보드와 통계 화면에 표시된다.",
        fonts, fills, borders, alignments
    )
    
    row = add_usecase_detail_table(
        ws_ucs, row, "UC-11", "훈련 데이터 시각화 분석",
        "일반 회원",
        "사용자가 로그인 상태여야 한다.",
        "1) 사용자가 통계 분석 메뉴로 진입한다.\n"
        "2) 시스템은 사용자의 RUN_RECORD 데이터를 기반으로 주간/월간/연간 누적 주행 거리를 바/선 그래프로 렌더링한다.\n"
        "3) 훈련 유형 분포(LSD, 인터벌 등), 온도 및 날씨 기상 환경에 따른 페이스의 변화율 상관 분석 데이터를 통계 수치로 표현한다.\n"
        "4) 전월 및 전년 동월 대비 주행 거리, 페이스, 평균 심박수를 대조할 수 있는 비교 통계 분석 테이블을 출력한다.",
        "2a) 훈련 기록 데이터가 부족(예: 3건 미만)하여 그래프 생성이 어려운 경우:\n"
        "    - '충분한 분석 데이터를 위해 기록을 더 등록해 주세요' 안내 메시지를 출력한다.",
        "데이터 분석 시각화 차트가 화면에 표시된다.",
        fonts, fills, borders, alignments
    )
    
    add_usecase_detail_table(
        ws_ucs, row, "UC-12", "동물 캐릭터 레벨 조회",
        "일반 회원",
        "사용자가 로그인 상태여야 한다.",
        "1) 사용자가 대시보드 또는 마이페이지에 접근한다.\n"
        "2) 시스템은 사용자의 전체 RUN_RECORD 주행 거리(Total Distance)를 합산한다.\n"
        "3) 합산 거리에 매핑된 동물 등급을 판별한다.\n"
        "   - 예: 100km 미만 (고양이), 100~300km (사슴), 300~600km (늑대), 600km 이상 (치타 러너)\n"
        "4) 등급에 해당하는 동물 캐릭터 이미지와 남은 레벨업 잔여 거리를 게이지 바 형태로 시각화하여 노출한다.",
        "2a) 누적 거리가 0km 인 경우:\n"
        "    - 최하위 등급인 '고양이 러너' 캐릭터와 함께 첫 러닝 등록 권장 가이드를 띄운다.",
        "사용자의 누적 거리에 맞춘 동물 등급과 캐릭터 정보가 출력된다.",
        fonts, fills, borders, alignments
    )

    # ==========================================
    # 4. Tab: 기능 요구사항 (FR)
    # ==========================================
    ws_fr = wb.create_sheet("기능 요구사항 (FR)")
    ws_fr.views.sheetView[0].showGridLines = True
    ws_fr.column_dimensions['A'].width = 3
    ws_fr.column_dimensions['B'].width = 18
    ws_fr.column_dimensions['C'].width = 28
    ws_fr.column_dimensions['D'].width = 12
    ws_fr.column_dimensions['E'].width = 65
    
    ws_fr.row_dimensions[2].height = 36
    ws_fr.merge_cells("B2:E2")
    t_cell = ws_fr["B2"]
    t_cell.value = "2.3 기능 요구사항 명세표 (Functional Requirements)"
    t_cell.font = fonts['sheet_title']
    t_cell.alignment = alignments['left_pad']
    for col in range(2, 6):
        ws_fr.cell(row=2, column=col).border = borders['heavy']
        
    ws_fr.row_dimensions[4].height = 24
    for idx, h in enumerate(["요구사항 ID", "요구사항명", "우선순위", "상세 설명"], 2):
        c = ws_fr.cell(row=4, column=idx, value=h)
        c.font = fonts['header']
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    fr_data = [
        ("FR-01", "소셜 로그인 연동", "필수", "구글 및 네이버 소셜 계정 로그인 연동(OAuth2 프로토콜 규격 사용)"),
        ("FR-02", "러닝 기록 등록 및 기상 연동", "필수", "날짜, 시간, 거리, 페이스, 심박, RPE(운동강도 1~10) 및 기상 API 연동 환경 정보 등록"),
        ("FR-03", "러닝 기록 수정 및 삭제", "필수", "작성자 본인이 등록한 훈련 데이터에 한해 수정 및 삭제 기능 제공 (소유권 유효성 검증)"),
        ("FR-04", "일별/주간/월간 기록 조회", "필수", "기록 리스트 페이징 조회 및 주간/월간 단위 누적 거리 및 평균 페이스 합산 조회"),
        ("FR-05", "월간 목표 및 대회 타겟 설정", "필수", "연/월별 정량적 주행 목표 설정 및 5k/10k/하프/풀코스 타겟 페이스 계획 관리"),
        ("FR-06", "마라톤 대회 및 기록증 보관", "필수", "대회 기본 스케줄 등록 및 증빙용 공식 기록증 이미지 업로드 기능 (최대 10MB)"),
        ("FR-07", "개인 최고 기록(PB) 자동 추출", "필수", "기록 데이터 연산을 통해 거리별 최고 완주 속도(PB) 상시 자동 갱신 및 표출"),
        ("FR-08", "대시보드 통계 그래프 시각화", "보통", "주간, 월간, 연간 누적 거리 추이 선/바 차트 및 전월/전년 동월 대비 분석 통계 시각화"),
        ("FR-09", "날씨 조건별 페이스 영향도 분석", "낮음", "기온, 습도, 날씨 맑음/비 상태에 따른 평균 페이스 영향도 및 비교 통계 분석"),
        ("FR-10", "캐릭터 기반 러닝 레벨 시스템", "낮음", "사용자의 전체 누적 거리에 맞추어 동물 캐릭터 등급(고양이~치타) 자동 부여 및 조회")
    ]
    for r_idx, record in enumerate(fr_data, 5):
        ws_fr.row_dimensions[r_idx].height = 20
        row_fill = fills['zebra'] if r_idx % 2 == 0 else fills['white']
        req_id, req_name, priority, desc = record
        
        c_id = ws_fr.cell(row=r_idx, column=2, value=req_id)
        c_name = ws_fr.cell(row=r_idx, column=3, value=req_name)
        c_prio = ws_fr.cell(row=r_idx, column=4, value=priority)
        c_desc = ws_fr.cell(row=r_idx, column=5, value=desc)
        
        for c in (c_id, c_name, c_prio, c_desc):
            c.font = fonts['data']
            c.fill = row_fill
            c.border = borders['thin']
            
        c_id.alignment = alignments['center']
        c_name.alignment = alignments['left_pad']
        c_prio.alignment = alignments['center']
        c_desc.alignment = alignments['left_pad']
        
        if priority == "필수":
            c_prio.fill = fill_essential
            c_prio.font = font_essential
        elif priority == "보통":
            c_prio.fill = fill_normal
            c_prio.font = font_normal
        elif priority == "낮음":
            c_prio.fill = fill_low
            c_prio.font = font_low

    # ==========================================
    # 5. Tab: 비기능 요구사항 (NFR)
    # ==========================================
    ws_nfr = wb.create_sheet("비기능 요구사항 (NFR)")
    ws_nfr.views.sheetView[0].showGridLines = True
    ws_nfr.column_dimensions['A'].width = 3
    ws_nfr.column_dimensions['B'].width = 18
    ws_nfr.column_dimensions['C'].width = 20
    ws_nfr.column_dimensions['D'].width = 75
    
    ws_nfr.row_dimensions[2].height = 36
    ws_nfr.merge_cells("B2:D2")
    t_cell = ws_nfr["B2"]
    t_cell.value = "2.4 비기능 요구사항 명세표 (Non-Functional Requirements)"
    t_cell.font = fonts['sheet_title']
    t_cell.alignment = alignments['left_pad']
    for col in range(2, 5):
        ws_nfr.cell(row=2, column=col).border = borders['heavy']
        
    ws_nfr.row_dimensions[4].height = 24
    for idx, h in enumerate(["요구사항 ID", "구분", "요구사항 내용"], 2):
        c = ws_nfr.cell(row=4, column=idx, value=h)
        c.font = fonts['header']
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    nfr_data = [
        ("NFR-01", "보안 (인증)", "OAuth2 소셜 연동을 단일 가입 방식으로 지원하며, Spring Security 설정을 통해 인가 프로세스를 관리한다."),
        ("NFR-02", "보안 (권한)", "로그인하지 않은 외부 방문자의 개인 기록 조회 및 이미지 파일 URI 직접 접근을 서버 단에서 거부(Forbidden) 처리한다."),
        ("NFR-03", "데이터 무결성", "회원이 탈퇴하는 경우, 해당 사용자가 업로드한 기록증 이미지 파일과 모든 훈련 기록 데이터를 연쇄적으로 제거(CASCADE)한다."),
        ("NFR-04", "성능 (대시보드)", "사용자별 훈련 목록 및 대시보드 메인 페이지의 조회 최적화를 위해 데이터베이스 인덱싱을 적용하고, 평균 응답 1.5초를 유지한다."),
        ("NFR-05", "호환성 (데이터베이스)", "MyBatis SQL 매퍼 쿼리 작성 시 표준 ANSI SQL을 준수하여 MySQL 8 및 Oracle DB 다중 데이터베이스 상에서 원활히 동작하도록 보장한다."),
        ("NFR-06", "사용성 (반응형 UI)", "데스크톱 브라우저뿐만 아니라 스마트폰 브라우저 화면 크기에 맞추어 레이아웃이 자동 조정되도록 반응형 그리드를 구현한다."),
        ("NFR-07", "확장성 (가민 워치 연동)", "향후 GPS 가민 스마트 워치 기기에서 추출되는 FIT 파일의 업로드 및 파싱 모듈 연동이 용이하도록 계층적 결합도를 낮추어 설계한다."),
        ("NFR-08", "오류 예외 처리", "서버 비즈니스 오류 발생 시 사용자용 커스텀 응답 구조(Error Response DTO) 및 HTTP Status Code를 일관성 있게 규격화하여 전달한다."),
        ("NFR-09", "확장성 (AI 모델)", "추후 v3의 AI 러닝 코칭(목표 가능성 분석 등) 학습을 대비해 일일 기록 RPE 및 환경 정보의 정형 데이터 규격을 통일하여 보관한다."),
        ("NFR-10", "운영 (로깅)", "SLF4J/Logback 프레임워크를 기반으로 주요 데이터 변경 트랜잭션, 파일 업로드 실패, 인증 예외 상황에 대한 추적 로그를 유지한다.")
    ]
    for r_idx, record in enumerate(nfr_data, 5):
        ws_nfr.row_dimensions[r_idx].height = 20
        row_fill = fills['zebra'] if r_idx % 2 == 0 else fills['white']
        for c_idx, val in enumerate(record, 2):
            c = ws_nfr.cell(row=r_idx, column=c_idx, value=val)
            c.font = fonts['data']
            c.fill = row_fill
            c.border = borders['thin']
            if c_idx in (2, 3):
                c.alignment = alignments['center']
                if c_idx == 3:
                    c.font = Font(name=font_name, size=9.5, bold=True, color="1A365D")
            else:
                c.alignment = alignments['left_pad']

    filename = "C:/Runners_plan/산출물/요구사항_명세서_Running_Coach_v1.2.xlsx"
    wb.save(filename)
    print(f"Successfully created Excel Document with Usecase details '{filename}'")

if __name__ == "__main__":
    main()
