import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    wb = openpyxl.Workbook()
    
    # Remove default sheet to build custom order
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Shared Style Declarations
    font_name = "Malgun Gothic"
    title_font = Font(name=font_name, size=14, bold=True, color="1A365D")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=9.5, color="2D3748")
    table_title_font = Font(name=font_name, size=10, bold=True, color="1A365D")
    
    fills = {
        'header': PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid"),
        'label': PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid"),
        'white': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        'zebra': PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    }
    alignments = {
        'center': Alignment(horizontal="center", vertical="center"),
        'left_pad': Alignment(horizontal="left", vertical="center", indent=1),
        'left_wrap': Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    }
    borders = {
        'thin': Border(left=Side(style='thin', color='CBD5E0'), right=Side(style='thin', color='CBD5E0'),
                       top=Side(style='thin', color='CBD5E0'), bottom=Side(style='thin', color='CBD5E0')),
        'heavy': Border(bottom=Side(style='medium', color='1A365D'))
    }

    # ==========================================
    # 1. Tab: 버전관리 (Version History)
    # ==========================================
    ws_ver = wb.create_sheet("버전관리")
    ws_ver.views.sheetView[0].showGridLines = True
    ws_ver.column_dimensions['A'].width = 3
    ws_ver.column_dimensions['B'].width = 12
    ws_ver.column_dimensions['C'].width = 16
    ws_ver.column_dimensions['D'].width = 48
    ws_ver.column_dimensions['E'].width = 22
    ws_ver.column_dimensions['F'].width = 15
    
    ws_ver.row_dimensions[2].height = 36
    ws_ver.merge_cells("B2:F2")
    t_cell = ws_ver["B2"]
    t_cell.value = "인덱스 설계서 - 버전 관리 이력"
    t_cell.font = title_font
    t_cell.alignment = alignments['left_pad']
    for col in range(2, 7):
        ws_ver.cell(row=2, column=col).border = borders['heavy']
        
    ws_ver.row_dimensions[4].height = 24
    for idx, h in enumerate(["버전", "변경 일자", "변경 내용", "작성자", "비고"], 2):
        c = ws_ver.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    ver_records = [
        ("v1.0", "2026-06-23", "테이블 정보 및 화면 설계를 반영한 최초 인덱스 설계서 작성", "Running Coach DB 기획팀", "-")
    ]
    for r_idx, record in enumerate(ver_records, 5):
        ws_ver.row_dimensions[r_idx].height = 20
        row_fill = fills['zebra'] if r_idx % 2 == 0 else fills['white']
        for c_idx, val in enumerate(record, 2):
            c = ws_ver.cell(row=r_idx, column=c_idx, value=val)
            c.font = data_font
            c.fill = row_fill
            c.border = borders['thin']
            if c_idx in (2, 3, 5):
                c.alignment = alignments['center']
            else:
                c.alignment = alignments['left_pad']

    # ==========================================
    # 2. Tab: 인덱스 설계서 (Index Specification)
    # ==========================================
    ws_idx = wb.create_sheet("인덱스 설계서")
    ws_idx.views.sheetView[0].showGridLines = True
    ws_idx.column_dimensions['A'].width = 3
    ws_idx.column_dimensions['B'].width = 8   # No
    ws_idx.column_dimensions['C'].width = 20  # 대상 테이블
    ws_idx.column_dimensions['D'].width = 26  # 인덱스명
    ws_idx.column_dimensions['E'].width = 28  # 인덱스 컬럼
    ws_idx.column_dimensions['F'].width = 12  # 고유 여부
    ws_idx.column_dimensions['G'].width = 85  # 선정 사유 및 연관 화면 (Rationale)
    
    # Sheet Title
    ws_idx.row_dimensions[2].height = 36
    ws_idx.merge_cells("B2:G2")
    title_cell = ws_idx["B2"]
    title_cell.value = "프로젝트 데이터베이스 인덱스 설계서"
    title_cell.font = title_font
    title_cell.alignment = alignments['left_pad']
    for col in range(2, 8):
        ws_idx.cell(row=2, column=col).border = borders['heavy']
        
    # Table Headers
    ws_idx.row_dimensions[4].height = 24
    idx_headers = ["No", "대상 테이블", "인덱스명", "인덱스 컬럼", "고유 여부", "선정 사유 및 연관 화면 (화면/기능 매핑)"]
    for idx, h in enumerate(idx_headers, 2):
        c = ws_idx.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    # Index Design Data (Based on tables & screens)
    indexes = [
        ("1", "RUN_RECORD", "IDX_RUN_RECORD_01", "runner_id, run_date", "Non-Unique", 
         "[연관 화면: 대시보드 - 최근 운동 기록 / 러닝 기록 - 일별 조회]\n"
         "회원별 러닝 기록 조회 시, 최신 날짜 역순으로 데이터를 정렬하여 리스트를 렌더링하는 액세스 빈도가 매우 높음.\n"
         "runner_id와 run_date를 묶은 복합 인덱스로 구성하여 데이터 조회 필터링 및 날짜 기준 정렬 연산 성능을 최적화함."),
         
        ("2", "RUN_RECORD", "IDX_RUN_RECORD_02", "training_type_code", "Non-Unique", 
         "[연관 화면: 통계 분석 - 훈련 유형 분석]\n"
         "러너의 훈련 데이터 중 LSD, 인터벌, 템포런 등의 훈련 비중 및 분포 통계를 도출하기 위해 training_type_code 기준으로 그룹바이(GROUP BY) 집계 쿼리가 빈번히 발생함. 그룹핑 성능 향상을 위해 인덱스 지정."),
         
        ("3", "GOAL", "IDX_GOAL_01", "runner_id", "Non-Unique", 
         "[연관 화면: 대시보드 - 이번 달 목표 및 달성률 / 목표 관리 - 월간 및 레이스 목표]\n"
         "로그인 회원별 당월 누적 거리 목표, 횟수 목표, 그리고 거리별 타겟 완주 시간 목표 데이터 조회 시 사용됨.\n"
         "사용자 ID(runner_id) 조건 필터링 성능을 확보하여 실시간 대시보드 렌더링 시 응답 지연을 방지함."),
         
        ("4", "RACE_RECORD", "IDX_RACE_RECORD_01", "runner_id", "Non-Unique", 
         "[연관 화면: 대시보드 - 다가오는 레이스 및 PB 현황 / 대회 기록 관리 / 통계 - PB 분석]\n"
         "사용자별 참여 마라톤 일정 목록, D-Day 디데이 연산, 그리고 개인 최고 기록(Personal Best) 자동 수집을 위해 RUNNER 테이블과 조인 및 runner_id 단위 데이터 조회 쿼리를 고속화함.")
    ]
    
    current_row = 5
    for r_idx, record in enumerate(indexes):
        # Calculate row height dynamically based on description lines
        lines = record[5].count('\n') + 1
        ws_idx.row_dimensions[current_row].height = max(24, lines * 18)
        
        row_fill = fills['zebra'] if r_idx % 2 == 1 else fills['white']
        
        for c_idx, val in enumerate(record, 2):
            c = ws_idx.cell(row=current_row, column=c_idx, value=val)
            c.font = data_font
            c.fill = row_fill
            c.border = borders['thin']
            
            # Formatting and Alignment
            if c_idx in (2, 3, 4, 5, 6): # No, Table, Index Name, Columns, Uniqueness
                c.alignment = alignments['center']
                if c_idx == 3: # Target Table
                    c.font = table_title_font
                elif c_idx == 4: # Index Name
                    c.font = Font(name=font_name, size=9.5, bold=True, color="2B6CB0")
            else: # Rationale/Screen Mapping
                c.alignment = alignments['left_wrap']
                
        current_row += 1

    filename = "c:/Runners_plan/산출물/인덱스_설계서_Running_Coach_v1.0.xlsx"
    wb.save(filename)
    print(f"Successfully created Index Specification Excel '{filename}'")

if __name__ == "__main__":
    main()
