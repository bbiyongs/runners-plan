import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    wb = openpyxl.Workbook()
    
    # Remove default sheet to build custom order
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Shared Style Declarations
    font_name = "Malgun Gothic"
    title_font = Font(name=font_name, size=14, bold=True, color="1A365D")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=9.5, color="2D3748")
    table_title_font = Font(name=font_name, size=10, bold=True, color="1A365D")
    
    fills = {
        'header': PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid"),
        'label': PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid"),
        'white': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        'zebra': PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    }
    alignments = {
        'center': Alignment(horizontal="center", vertical="center"),
        'left_pad': Alignment(horizontal="left", vertical="center", indent=1),
        'left_wrap': Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    }
    borders = {
        'thin': Border(left=Side(style='thin', color='CBD5E0'), right=Side(style='thin', color='CBD5E0'),
                       top=Side(style='thin', color='CBD5E0'), bottom=Side(style='thin', color='CBD5E0')),
        'heavy': Border(bottom=Side(style='medium', color='1A365D'))
    }

    # ==========================================
    # 1. Tab: 버전관리 (Version History)
    # ==========================================
    ws_ver = wb.create_sheet("버전관리")
    ws_ver.views.sheetView[0].showGridLines = True
    ws_ver.column_dimensions['A'].width = 3
    ws_ver.column_dimensions['B'].width = 12
    ws_ver.column_dimensions['C'].width = 16
    ws_ver.column_dimensions['D'].width = 48
    ws_ver.column_dimensions['E'].width = 22
    ws_ver.column_dimensions['F'].width = 15
    
    ws_ver.row_dimensions[2].height = 36
    ws_ver.merge_cells("B2:F2")
    t_cell = ws_ver["B2"]
    t_cell.value = "인덱스 설계서 - 버전 관리 이력"
    t_cell.font = title_font
    t_cell.alignment = alignments['left_pad']
    for col in range(2, 7):
        ws_ver.cell(row=2, column=col).border = borders['heavy']
        
    ws_ver.row_dimensions[4].height = 24
    for idx, h in enumerate(["버전", "변경 일자", "변경 내용", "작성자", "비고"], 2):
        c = ws_ver.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    ver_records = [
        ("v1.0", "2026-06-23", "테이블 정보 및 화면 설계를 반영한 최초 인덱스 설계서 작성", "Running Coach DB 기획팀", "-"),
        ("v1.1", "2026-06-23", "테이블 명세서 v1.1 기준 인덱스 고도화 (Unique 제약 및 FK 인덱스 4종 신규 반영)", "Running Coach DB 기획팀", "-"),
        ("v1.2", "2026-06-24", "인덱스 6종 신규 및 변경 반영 (복합 인덱스 고도화, 외래키 및 Unique 제약 보완)", "Running Coach DB 기획팀", "-"),
        ("v1.3", "2026-06-24", "RUNNER 테이블의 email 컬럼 삭제에 따른 UIDX_RUNNER_01 인덱스 제거", "Running Coach DB 기획팀", "-")
    ]
    for r_idx, record in enumerate(ver_records, 5):
        ws_ver.row_dimensions[r_idx].height = 20
        row_fill = fills['zebra'] if r_idx % 2 == 0 else fills['white']
        for c_idx, val in enumerate(record, 2):
            c = ws_ver.cell(row=r_idx, column=c_idx, value=val)
            c.font = data_font
            c.fill = row_fill
            c.border = borders['thin']
            if c_idx in (2, 3, 5):
                c.alignment = alignments['center']
            else:
                c.alignment = alignments['left_pad']

    # ==========================================
    # 2. Tab: 인덱스 설계서 (Index Specification)
    # ==========================================
    ws_idx = wb.create_sheet("인덱스 설계서")
    ws_idx.views.sheetView[0].showGridLines = True
    ws_idx.column_dimensions['A'].width = 3
    ws_idx.column_dimensions['B'].width = 8   # No
    ws_idx.column_dimensions['C'].width = 24  # 대상 테이블
    ws_idx.column_dimensions['D'].width = 28  # 인덱스명
    ws_idx.column_dimensions['E'].width = 36  # 인덱스 컬럼
    ws_idx.column_dimensions['F'].width = 12  # 고유 여부
    ws_idx.column_dimensions['G'].width = 85  # 선정 사유 및 연관 화면 (Rationale)
    
    # Sheet Title
    ws_idx.row_dimensions[2].height = 36
    ws_idx.merge_cells("B2:G2")
    title_cell = ws_idx["B2"]
    title_cell.value = "프로젝트 데이터베이스 인덱스 설계서 (v1.3)"
    title_cell.font = title_font
    title_cell.alignment = alignments['left_pad']
    for col in range(2, 8):
        ws_idx.cell(row=2, column=col).border = borders['heavy']
        
    # Table Headers
    ws_idx.row_dimensions[4].height = 24
    idx_headers = ["No", "대상 테이블", "인덱스명", "인덱스 컬럼", "고유 여부", "선정 사유 및 연관 화면 (화면/기능 매핑)"]
    for idx, h in enumerate(idx_headers, 2):
        c = ws_idx.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = fills['header']
        c.alignment = alignments['center']
        c.border = borders['thin']
        
    # Updated Index Design Data (11 indexes, email index removed)
    indexes = [
        ("1", "RUNNER_SOCIAL_ACCOUNT", "UIDX_RUNNER_SOCIAL_ACCOUNT_01", "provider, provider_user_id", "Unique", 
         "[연관 화면: 소셜 로그인]\n"
         "동일 소셜 계정이 여러 러너 계정에 중복 연동되는 것을 차단하는 데이터 무결성(Unique) 제약 조건.\n"
         "로그인 요청 시 소셜 제공자 정보 기반 사용자 식별 및 토큰 검증 쿼리 성능 고속화."),
         
        ("2", "RUNNER_SOCIAL_ACCOUNT", "IDX_RUNNER_SOCIAL_ACCOUNT_01", "runner_id", "Non-Unique", 
         "[연관 화면: 마이페이지 - 연동 소셜 계정 확인]\n"
         "회원별 소셜 계정 연동 리스트 조회 성능 확보 및 회원 탈퇴 시 외래키 참조 무결성 검증(Delete CASCADE) 속도 향상을 위한 외래키(FK) 인덱스."),
         
        ("3", "RUN_RECORD", "IDX_RUN_RECORD_01", "runner_id, run_date", "Non-Unique", 
         "[연관 화면: 대시보드 - 최근 운동 기록 / 러닝 기록 - 일별 조회]\n"
         "회원별 러닝 기록 조회 시 최신 날짜 역순으로 데이터를 정렬하여 리스트를 렌더링하는 액세스 빈도가 매우 높음.\n"
         "runner_id와 run_date를 묶은 복합 인덱스로 구성하여 데이터 조회 필터링 및 날짜 기준 정렬 연산 성능을 최적화함."),
         
        ("4", "RUN_RECORD", "IDX_RUN_RECORD_02", "runner_id, training_type_code", "Non-Unique", 
         "[연관 화면: 통계 분석 - 훈련 유형 분석]\n"
         "회원별 특정 훈련 유형(LSD, 인터벌, 템포런 등)의 통계 집계 및 분석 조회 성능 향상을 위한 복합 인덱스.\n"
         "runner_id 조건 필터링 후 training_type_code 그룹바이(GROUP BY) 집계 쿼리의 속도를 대폭 최적화함."),
         
        ("5", "RUN_RECORD", "IDX_RUN_RECORD_03", "run_date", "Non-Unique", 
         "[연관 기능: 전체 통계 및 기간별 조회]\n"
         "개별 사용자에 국한되지 않고 특정 일자나 전체 기간 조건으로 훈련 데이터를 필터링하거나 통계를 집계할 때의 쿼리 성능 향상을 위한 날짜 단일 인덱스."),
         
        ("6", "RUN_RECORD", "IDX_RUN_RECORD_04", "runner_id, run_datetime", "Non-Unique", 
         "[연관 화면: 대시보드 - 최근 운동 기록]\n"
         "사용자별 최근 러닝 일시(run_datetime) 기준 내림차순 정렬 조회가 매우 빈번함.\n"
         "초 단위 일시 정보를 포함한 정렬 연산의 부하를 제거하고, 최신 N건 조회 쿼리 속도를 극대화하기 위해 runner_id와 run_datetime을 결합한 복합 인덱스."),
         
        ("7", "GOAL", "UIDX_GOAL_01", "runner_id, target_year, target_month, goal_type_code", "Unique", 
         "[연관 화면: 대시보드 - 이번 달 목표 및 달성률 / 목표 관리 - 월간 목표]\n"
         "동일 러너가 특정 월에 동일한 유형의 목표(예: 2026년 7월 월간 거리 목표)를 중복 등록하는 오작동을 DB 제약으로 원천 방지.\n"
         "사용자 ID(runner_id) 조건 및 년/월 필터링을 동시 수행하므로 단일 컬럼 인덱스 대비 월등한 검색 속도 보장."),
         
        ("8", "RACE_RECORD", "IDX_RACE_RECORD_01", "runner_id", "Non-Unique", 
         "[연관 화면: 대시보드 - 다가오는 레이스 및 PB 현황 / 대회 기록 관리 / 통계 - PB 분석]\n"
         "사용자별 참여 마라톤 일정 목록, D-Day 디데이 연산, 그리고 개인 최고 기록(Personal Best) 자동 수집을 위해 RUNNER 테이블과 조인 및 runner_id 단위 데이터 조회 쿼리를 고속화함."),
         
        ("9", "RACE_RECORD", "IDX_RACE_RECORD_02", "runner_id, race_type_code", "Non-Unique", 
         "[연관 화면: 통계 - PB 분석 / 대회 기록 목록 조회]\n"
         "사용자별 대회 완주 기록 중 특정 대회 유형(5km, 10km, 하프, 풀코스 등)별 기록과 개인 최고 기록(PB) 데이터를 빠르게 분석 및 정렬하기 위한 복합 인덱스."),
         
        ("10", "CODE_DETAIL", "UIDX_CODE_DETAIL_01", "group_code, code_value", "Unique", 
         "[연관 기능: 공통 코드 무결성 확보]\n"
         "특정 그룹코드(group_code) 내에서 상세 코드값(code_value)이 중복되어 등록되는 데이터 오류를 방지하고 코드 체계의 일관성과 무결성을 DB 레벨에서 강제하기 위한 복합 Unique 인덱스."),
         
        ("11", "CODE_DETAIL", "IDX_CODE_DETAIL_02", "group_code", "Non-Unique", 
         "[연관 기능: 공통 코드 상세 매핑]\n"
         "코드 테이블에서 특정 그룹코드의 코드 상세 리스트를 빈번히 룩업 조회할 때 성능 향상.\n"
         "CODE_GROUP 테이블과의 조인 연산 처리 및 데이터 삭제 시 FK 제약조건 체크 속도를 높이기 위한 외래키 인덱스.")
    ]
    
    current_row = 5
    for r_idx, record in enumerate(indexes):
        lines = record[5].count('\n') + 1
        ws_idx.row_dimensions[current_row].height = max(24, lines * 18)
        
        row_fill = fills['zebra'] if r_idx % 2 == 1 else fills['white']
        
        for c_idx, val in enumerate(record, 2):
            c = ws_idx.cell(row=current_row, column=c_idx, value=val)
            c.font = data_font
            c.fill = row_fill
            c.border = borders['thin']
            
            # Formatting and Alignment
            if c_idx in (2, 3, 4, 5, 6): # No, Table, Index Name, Columns, Uniqueness
                c.alignment = alignments['center']
                if c_idx == 3: # Target Table
                    c.font = table_title_font
                elif c_idx == 4: # Index Name
                    c.font = Font(name=font_name, size=9.5, bold=True, color="2B6CB0")
                elif c_idx == 6: # Uniqueness
                    if val == "Unique":
                        c.font = Font(name=font_name, size=9.5, bold=True, color="9B2C2C")
            else: # Rationale/Screen Mapping
                c.alignment = alignments['left_wrap']
                
        current_row += 1

    filename = "c:/Runners_plan/산출물/인덱스_설계서_Running_Coach_v1.1.xlsx"
    wb.save(filename)
    print(f"Successfully created Index Specification Excel '{filename}'")

if __name__ == "__main__":
    main()
