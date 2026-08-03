import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    wb = openpyxl.Workbook()
    
    # Remove default sheet to build custom order
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Fonts
    font_name = "Malgun Gothic"
    title_font = Font(name=font_name, size=15, bold=True, color="1A365D")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10, color="2D3748")
    
    # Fills
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Priority colors
    fill_high = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
    font_high = Font(name=font_name, size=10, bold=True, color="9B2C2C")
    
    fill_med = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")
    font_med = Font(name=font_name, size=10, bold=True, color="975A16")
    
    fill_low = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    font_low = Font(name=font_name, size=10, bold=True, color="4A5568")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Borders
    thin_border_side = Side(style='thin', color='CBD5E0')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    bottom_heavy_border = Border(bottom=Side(style='medium', color='1A365D'))

    # ==========================================
    # 1. Tab: 버전관리 (Version History)
    # ==========================================
    ws_ver = wb.create_sheet("버전관리")
    ws_ver.views.sheetView[0].showGridLines = True
    
    # Column Widths
    ws_ver.column_dimensions['A'].width = 3
    ws_ver.column_dimensions['B'].width = 12  # 버전
    ws_ver.column_dimensions['C'].width = 16  # 변경일자
    ws_ver.column_dimensions['D'].width = 45  # 변경 내용
    ws_ver.column_dimensions['E'].width = 22  # 작성자
    ws_ver.column_dimensions['F'].width = 15  # 비고
    
    # Title
    ws_ver.row_dimensions[2].height = 40
    ws_ver.merge_cells("B2:F2")
    t_cell = ws_ver["B2"]
    t_cell.value = "요구사항 명세서 - 버전 관리 이력"
    t_cell.font = title_font
    t_cell.alignment = align_left
    for col in range(2, 7):
        ws_ver.cell(row=2, column=col).border = bottom_heavy_border
        
    # Table Headers
    ws_ver.row_dimensions[4].height = 24
    ver_headers = ["버전", "변경 일자", "변경 내용", "작성자", "비고"]
    for idx, h in enumerate(ver_headers, 2):
        c = ws_ver.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = thin_border
        
    # Initial Version Data
    ws_ver.row_dimensions[5].height = 22
    ver_data = ["v1.0", "2026-06-18", "최초 요구사항 정의서 작성 및 멀티 탭 문서화", "Running Coach 개발팀", "-"]
    for idx, val in enumerate(ver_data, 2):
        c = ws_ver.cell(row=5, column=idx, value=val)
        c.font = data_font
        c.border = thin_border
        if idx in (2, 3, 5):  # Version, Date, Author
            c.alignment = align_center
        else:
            c.alignment = align_left
            if idx == 4:  # Details
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ==========================================
    # 2. Tabs for Categories
    # ==========================================
    categories = {
        "회원": [
            ("USER-001", "구글 로그인", "구글 계정 로그인", "상"),
            ("USER-002", "네이버 로그인", "네이버 계정 로그인", "상"),
            ("USER-003", "프로필 조회", "사용자 정보 조회", "중"),
            ("USER-004", "프로필 수정", "닉네임 수정", "중"),
        ],
        "러닝 기록": [
            ("RUN-001", "기록 등록", "러닝 기록 등록", "상"),
            ("RUN-002", "기록 수정", "러닝 기록 수정", "상"),
            ("RUN-003", "기록 삭제", "러닝 기록 삭제", "상"),
            ("RUN-004", "일별 조회", "일별 기록 조회", "상"),
            ("RUN-005", "주간 조회", "주간 기록 조회", "중"),
            ("RUN-006", "월간 조회", "월간 기록 조회", "중"),
        ],
        "목표관리": [
            ("GOAL-001", "월간 목표", "거리 목표 설정", "상"),
            ("GOAL-002", "월간 목표", "운동 횟수 목표 설정", "중"),
            ("GOAL-003", "10km 목표", "목표 기록 설정", "중"),
            ("GOAL-004", "하프 목표", "목표 기록 설정", "중"),
            ("GOAL-005", "풀코스 목표", "목표 기록 설정", "중"),
        ],
        "통계": [
            ("STAT-001", "PB 조회", "개인 최고기록 조회", "중"),
            ("STAT-002", "월별 분석", "월별 거리 분석", "중"),
            ("STAT-003", "목표 달성률", "목표 대비 달성률 계산", "중"),
            ("STAT-004", "페이스 추이", "평균 페이스 변화", "하"),
            ("STAT-005", "러닝 레벨", "레벨 계산", "하"),
        ]
    }
    
    for cat_name, items in categories.items():
        ws = wb.create_sheet(cat_name)
        ws.views.sheetView[0].showGridLines = True
        
        # Column Widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 16  # 요구사항 ID
        ws.column_dimensions['C'].width = 24  # 기능명
        ws.column_dimensions['D'].width = 45  # 상세 설명
        ws.column_dimensions['E'].width = 12  # 우선순위
        
        # Title
        ws.row_dimensions[2].height = 40
        ws.merge_cells("B2:E2")
        title_cell = ws["B2"]
        title_cell.value = f"요구사항 명세서 - {cat_name} 기능"
        title_cell.font = title_font
        title_cell.alignment = align_left
        for col in range(2, 6):
            ws.cell(row=2, column=col).border = bottom_heavy_border
            
        # Table Headers
        ws.row_dimensions[4].height = 24
        headers = ["요구사항 ID", "기능명", "상세 설명", "우선순위"]
        for idx, h in enumerate(headers, 2):
            c = ws.cell(row=4, column=idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = align_center
            c.border = thin_border
            
        # Data Rows
        current_row = 5
        for r_idx, item in enumerate(items):
            ws.row_dimensions[current_row].height = 20
            row_fill = zebra_fill if r_idx % 2 == 1 else white_fill
            
            req_id, func_name, desc, priority = item
            
            c_id = ws.cell(row=current_row, column=2, value=req_id)
            c_func = ws.cell(row=current_row, column=3, value=func_name)
            c_desc = ws.cell(row=current_row, column=4, value=desc)
            c_prio = ws.cell(row=current_row, column=5, value=priority)
            
            # Format styles
            for c in (c_id, c_func, c_desc, c_prio):
                c.font = data_font
                c.fill = row_fill
                c.border = thin_border
                
            c_id.alignment = align_center
            c_func.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c_desc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c_prio.alignment = align_center
            
            # Priority color coding
            if priority == "상":
                c_prio.fill = fill_high
                c_prio.font = font_high
            elif priority == "중":
                c_prio.fill = fill_med
                c_prio.font = font_med
            elif priority == "하":
                c_prio.fill = fill_low
                c_prio.font = font_low
                
            current_row += 1
            
    filename = "요구사항_명세서_Running_Coach.xlsx"
    wb.save(filename)
    print(f"Successfully created Multi-Tab Excel: '{filename}'")

if __name__ == "__main__":
    main()
